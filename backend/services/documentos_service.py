# Archivo: backend/services/documentos_service.py
import io
import base64
import pandas as pd
from datetime import datetime
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.graphics.barcode import qr
from reportlab.graphics.shapes import Drawing
from reportlab.graphics import renderPDF
from openpyxl.styles import PatternFill, Font, Alignment

# ─── COLORES DE LA MARCA SOLARVER ───
BRAND_BLUE = colors.HexColor('#1E85C8')
BRAND_DARK = colors.HexColor('#0E4F8A')
BRAND_ORANGE = colors.HexColor('#FF7A1F')
BRAND_BG = colors.HexColor('#F0F4F8')

def pie_de_pagina(canvas, doc):
    """Dibuja el pie de página con el número de página dinámico."""
    canvas.saveState()
    canvas.setFont('Helvetica', 9)
    canvas.setFillColor(colors.gray)
    canvas.drawString(30, 20, "SolarVer - Sistema de Gestión")
    canvas.drawRightString(landscape(letter)[0] - 30, 20, f"Página {doc.page}")
    canvas.restoreState()

def generar_pdf_base64(cliente):
    """Genera el PDF del Estado de Cuenta mensual con diseño membretado y marcas de agua."""
    output = io.BytesIO()
    p = canvas.Canvas(output, pagesize=letter)
    width, height = letter
    
    # ── Marca de Agua Condicional ──
    estatus = str(cliente.get('Estatus', '')).lower()
    saldo = float(cliente.get('Saldo_Pendiente', 0))
    
    if saldo <= 0 or estatus == 'pagado':
        watermark_text = "P A G A D O"
        watermark_color = colors.Color(0, 0.8, 0, alpha=0.1) # Verde transparente
    elif estatus in ['vencido', 'atrasado', 'moroso']:
        watermark_text = "V E N C I D O"
        watermark_color = colors.Color(1, 0, 0, alpha=0.1) # Rojo transparente
    else:
        watermark_text = ""
        
    if watermark_text:
        p.saveState()
        p.setFillColor(watermark_color)
        p.translate(width/2, height/2)
        p.rotate(45)
        p.setFont("Helvetica-Bold", 80)
        p.drawCentredString(0, 0, watermark_text)
        p.restoreState()
    
    # ── Encabezado Gráfico ──
    p.setFillColor(BRAND_BLUE)
    p.rect(0, height - 75, width, 75, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 24)
    p.drawString(50, height - 48, "SolarVer")
    
    p.setFont("Helvetica", 13)
    p.drawString(width - 230, height - 45, "Estado de Cuenta Oficial")
    
    # ── Datos del Cliente ──
    p.setFillColor(colors.black)
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, height - 130, f"Cliente: {cliente['Cliente']}")
    
    p.setFont("Helvetica", 10)
    p.setFillColor(colors.darkgray)
    p.drawString(50, height - 150, f"Fecha de emisión: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    
    # ── Cuadro de Resumen ──
    p.setFillColor(BRAND_BG)
    p.roundRect(50, height - 280, width - 100, 100, 10, fill=1, stroke=0)
    
    p.setFillColor(BRAND_DARK)
    p.setFont("Helvetica-Bold", 12)
    p.drawString(70, height - 210, "Detalle de su financiamiento:")
    
    p.setFont("Helvetica", 11)
    p.setFillColor(colors.black)
    p.drawString(70, height - 235, f"Día de Corte: {cliente['Dia_Pago']} de cada mes")
    p.drawString(width/2 + 20, height - 235, f"Estatus actual: {cliente['Estatus'].upper()}")
    
    # Resaltar saldo dependiente de si hay deuda o no
    p.setFont("Helvetica-Bold", 14)
    p.setFillColor(colors.red if saldo > 0 else colors.green)
    p.drawString(70, height - 260, f"Saldo Pendiente: ${saldo:,.2f} MXN")
    
    # ── Pie de página ──
    p.setFillColor(colors.gray)
    p.setFont("Helvetica-Oblique", 10)
    p.drawCentredString(width / 2, 50, "Si ya realizó su pago, haga caso omiso a este documento.")
    
    p.showPage()
    p.save()
    pdf_bytes = output.getvalue()
    output.close()
    return base64.b64encode(pdf_bytes).decode('utf-8')

def generar_excel_reporte(datos):
    """Genera un archivo Excel con formato nativo (Filtros, Congelación y Moneda)."""
    df = pd.DataFrame(datos)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Reporte')
        worksheet = writer.sheets['Reporte']
        
        # Definir estilos de Excel
        header_fill = PatternFill(start_color="1E85C8", end_color="1E85C8", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Aplicar estilo a los encabezados
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        
        # Auto-ajustar ancho de columnas y aplicar formato de moneda nativo
        for column_cells in worksheet.columns:
            max_length = 0
            col_letter = column_cells[0].column_letter
            col_header = str(column_cells[0].value).lower()
            
            # Identificar columnas que deben tener formato de moneda
            es_moneda = any(palabra in col_header for palabra in ['monto', 'saldo', 'interés', 'interes'])
            
            for cell in column_cells:
                try:
                    # Aplicar formato de moneda a los datos (excluyendo el encabezado)
                    if cell.row > 1 and es_moneda:
                        cell.number_format = '"$"#,##0.00'
                        # Limpiar strings si vienen con símbolo de pesos desde la BD
                        if isinstance(cell.value, str):
                            cell.value = float(cell.value.replace('$', '').replace(',', ''))
                            
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Margen de holgura extra (+4)
            worksheet.column_dimensions[col_letter].width = min(max_length + 4, 50) 
            
        # Congelar la primera fila y activar auto-filtros
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
            
    output.seek(0)
    return output

def generar_pdf_reporte(datos, tipo):
    """Genera un reporte tabular en PDF con Zebra-Striping y paginación dinámica."""
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter),
                            rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    
    titulo_texto = "Reporte de Pagos Realizados" if tipo == 'realizados' else f"Reporte de Cobranza ({tipo.upper()})"
    
    # Título
    title_style = styles['Title']
    title_style.textColor = BRAND_DARK
    title_style.fontName = 'Helvetica-Bold'
    elements.append(Paragraph(f"SolarVer - {titulo_texto}", title_style))
    
    # Subtítulo (Periodo y Generación)
    sub_style = styles['Normal']
    sub_style.textColor = colors.gray
    sub_style.alignment = 1 # Centrado
    periodo = "Últimos 30 días" if tipo == 'realizados' else "Estado al corte actual"
    elements.append(Paragraph(f"Periodo: {periodo} &nbsp;&nbsp;|&nbsp;&nbsp; Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", sub_style))
    elements.append(Spacer(1, 25))
    
    # Construcción de la matriz de datos
    if tipo == 'realizados':
        tabla_data = [["Folio", "Cliente", "Contacto (Tel / Correo)", "Monto", "Método", "Fecha"]]
        for d in datos:
            contacto = f"{d['Telefono'] or '-'} \n{d['Correo'] or '-'}"
            tabla_data.append([
                d['Folio'], 
                str(d['Cliente'])[:30], 
                contacto, 
                f"${d['Monto']:,.2f}", 
                d['Metodo_Pago'], 
                str(d['Fecha_Pago'])
            ])
    else:
        tabla_data = [["Cliente", "Contacto (Tel / Correo)", "Día Pago", "Saldo", "Interés", "Estatus"]]
        for d in datos:
            contacto = f"{d['Telefono'] or '-'} \n{d['Correo'] or '-'}"
            tabla_data.append([
                str(d['Cliente'])[:30], 
                contacto, 
                f"Día {d['Dia_Pago']}", 
                f"${d['Saldo_Pendiente']:,.2f}", 
                f"${d['Interes_Acumulado']:,.2f}", 
                str(d['Estatus']).capitalize()
            ])
    
    # Dibujar la tabla
    t = Table(tabla_data, repeatRows=1)
    
    # Estilo base de la tabla
    estilo_tabla = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), BRAND_BLUE),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('PADDING', (0, 0), (-1, 0), 10),
        
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
    ])
    
    # Zebra-Striping (Intercalar color gris muy sutil)
    for i in range(1, len(tabla_data)):
        if i % 2 == 0:
            estilo_tabla.add('BACKGROUND', (0, i), (-1, i), BRAND_BG)
            
    t.setStyle(estilo_tabla)
    elements.append(t)
    
    # Construir documento con paginación
    doc.build(elements, onFirstPage=pie_de_pagina, onLaterPages=pie_de_pagina)
    
    output.seek(0)
    return output

def generar_pdf_instrucciones_pago(cliente_nombre, monto, clave_referencia, fecha_limite):
    """Genera un PDF con las instrucciones de pago e incluye un Código QR."""
    output = io.BytesIO()
    p = canvas.Canvas(output, pagesize=letter)
    width, height = letter
    
    # ── Encabezado Gráfico ──
    p.setFillColor(BRAND_BLUE)
    p.rect(0, height - 75, width, 75, fill=1, stroke=0)
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 24)
    p.drawString(50, height - 48, "SolarVer")
    p.setFont("Helvetica", 13)
    p.drawString(width - 210, height - 45, "Instrucciones de Pago")
    
    # ── Datos del cliente ──
    p.setFillColor(colors.black)
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, height - 130, f"Estimado/a: {cliente_nombre}")
    p.setFont("Helvetica", 12)
    p.drawString(50, height - 155, f"Fecha límite de pago: ")
    p.setFont("Helvetica-Bold", 12)
    p.setFillColor(BRAND_ORANGE)
    p.drawString(180, height - 155, f"{fecha_limite}")
    
    # ── Caja Principal ──
    p.setFillColor(BRAND_BG)
    p.roundRect(50, height - 440, width - 100, 250, 10, fill=1, stroke=0)
    
    p.setFillColor(BRAND_DARK)
    p.setFont("Helvetica-Bold", 14)
    p.drawString(70, height - 230, "Detalles para realizar su transferencia:")
    
    p.setFillColor(colors.black)
    p.setFont("Helvetica", 12)
    p.drawString(90, height - 265, "1. Ingrese a la aplicación de su banco.")
    p.drawString(90, height - 290, "2. Dé de alta la siguiente cuenta CLABE (Banco STP):")
    
    p.setFont("Helvetica-Bold", 12)
    p.drawString(110, height - 315, "CLABE: 646180123456789012")
    p.drawString(110, height - 335, "Beneficiario: SolarVer S.A. de C.V.")
    
    p.setFont("Helvetica", 12)
    p.drawString(90, height - 365, "3. Ingrese el monto exacto a pagar:")
    p.setFont("Helvetica-Bold", 14)
    p.setFillColor(colors.red)
    p.drawString(110, height - 390, f"${monto:,.2f} MXN")
    
    p.setFillColor(colors.black)
    p.setFont("Helvetica", 12)
    p.drawString(90, height - 420, "4. Es OBLIGATORIO colocar la siguiente referencia:")
    
    # Referencia destacada
    p.setFont("Helvetica-Bold", 18)
    p.setFillColor(BRAND_BLUE)
    p.drawCentredString(width / 2, height - 485, clave_referencia)
    
    # ── Código QR ──
    try:
        # Generar un código QR con los datos principales de la transferencia
        qr_data = f"CLABE:646180123456789012\nRef:{clave_referencia}\nMonto:{monto}"
        qr_code = qr.QrCodeWidget(qr_data)
        qr_code.barWidth = 90
        qr_code.barHeight = 90
        d = Drawing(90, 90)
        d.add(qr_code)
        renderPDF.draw(d, p, width - 160, height - 410)
    except Exception as e:
        print("Advertencia: No se pudo generar el QR", e)
    
    # Nota final
    p.setFillColor(colors.gray)
    p.setFont("Helvetica-Oblique", 10)
    p.drawCentredString(width / 2, height - 540, "Nota: Si no coloca la referencia exactamente como se muestra, su pago")
    p.drawCentredString(width / 2, height - 555, "no se registrará automáticamente y requerirá una conciliación manual.")
    
    p.showPage()
    p.save()
    
    pdf_bytes = output.getvalue()
    output.close()
    return base64.b64encode(pdf_bytes).decode('utf-8')